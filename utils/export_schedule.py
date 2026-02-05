"""
Schedule Export Utilities (Item 4)
Export schedules to CSV, Excel, and PDF formats.
"""
import json
from pathlib import Path
from io_handler import load_troops_from_json
from constrained_scheduler import ConstrainedScheduler
from activities import get_all_activities
from models import Day

def export_to_csv(schedule, troops, output_file):
    """Export schedule to CSV format."""
    import csv
    
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        
        # Header
        writer.writerow(['Troop', 'Day', 'Slot', 'Activity', 'Duration', 'Priority', 'Scouts', 'Adults'])
        
        # Sort entries by troop, then day, then slot
        sorted_entries = sorted(
            schedule.entries,
            key=lambda e: (e.troop.name, e.time_slot.day.value, e.time_slot.slot_number)
        )
       
        # Write data
        for entry in sorted_entries:
            # Skip continuation slots (already counted in duration)
            is_continuation = any(
                other.troop == entry.troop and 
                other.activity.name == entry.activity.name and
                other.time_slot.day == entry.time_slot.day and
                other.time_slot.slot_number < entry.time_slot.slot_number
                for other in schedule.entries
            )
            
            if not is_continuation:
                writer.writerow([
                    entry.troop.name,
                    entry.time_slot.day.value,
                    entry.time_slot.slot_number,
                    entry.activity.name,
                    entry.activity.slots,
                    entry.troop.get_priority(entry.activity.name),
                    entry.troop.scouts,
                    entry.troop.adults
                ])
    
    print(f"[OK] CSV exported to: {output_file}")

def export_to_excel(schedule, troops, output_file):
    """Export schedule to Excel format with formatted sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[WARN] openpyxl not installed. Run: pip install openpyxl")
        return
    
    wb = Workbook()
    
    # Sheet 1: Full Schedule
    ws = wb.active
    ws.title = "Full Schedule"
    
    # Header
    headers = ['Troop', 'Day', 'Slot', 'Activity', 'Duration', 'Priority', 'Scouts', 'Adults']
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    sorted_entries = sorted(
        schedule.entries,
        key=lambda e: (e.troop.name, e.time_slot.day.value, e.time_slot.slot_number)
    )
    
    for entry in sorted_entries:
        is_continuation = any(
            other.troop == entry.troop and 
            other.activity.name == entry.activity.name and
            other.time_slot.day == entry.time_slot.day and
            other.time_slot.slot_number < entry.time_slot.slot_number
            for other in schedule.entries
        )
        
        if not is_continuation:
            ws.append([
                entry.troop.name,
                entry.time_slot.day.value,
                entry.time_slot.slot_number,
                entry.activity.name,
                entry.activity.slots,
                entry.troop.get_priority(entry.activity.name),
                entry.troop.scouts,
                entry.troop.adults
            ])
    
    # Sheet 2: Troop Schedules
    ws2 = wb.create_sheet("Troop Schedules")
    
    # Create grid: Rows = Troops, Columns = Day-Slot pairs
    time_slots = [f"{day.value[:3]}-{slot}" for day in Day for slot in range(1, 4)]
    ws2.append(['Troop'] + time_slots)
    
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for troop in troops:
        row_data = [troop.name]
        for day in Day:
            for slot in range(1,  4):
                # Find activity for this troop/day/slot
                activities = [
                    e.activity.name for e in schedule.entries
                    if e.troop == troop and 
                       e.time_slot.day == day and 
                       e.time_slot.slot_number == slot
                ]
                row_data.append(activities[0] if activities else "")
        ws2.append(row_data)
    
    wb.save(output_file)
    print(f"[OK] Excel exported to: {output_file}")

def export_to_pdf(schedule, troops, output_file):
    """Export schedule to PDF format."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        print("[WARN] reportlab not installed. Run: pip install reportlab")
        return
    
    doc = SimpleDocTemplate(output_file, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    elements.append(Paragraph(f"<b>Summer Camp Schedule</b>", styles['Title']))
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Create a table for each troop
    for troop in troops:
        # Troop header
        elements.append(Paragraph(f"<b>{troop.name}</b> ({troop.scouts} scouts, {troop.adults} adults)", styles['Heading2']))
        
        # Create schedule grid
        data = [['Day/Slot', 'Slot 1', 'Slot 2', 'Slot 3']]
        
        for day in Day:
            row = [day.value]
            for slot in range(1, 4):
                activities = [
                    e.activity.name for e in schedule.entries
                    if e.troop == troop and 
                       e.time_slot.day == day and 
                       e.time_slot.slot_number == slot
                ]
                row.append(activities[0] if activities else "-")
            data.append(row)
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(PageBreak())
    
    doc.build(elements)
    print(f"[OK] PDF exported to: {output_file}")

def export_schedule(week_file, formats=['csv', 'excel', 'pdf']):
    """Export a week's schedule to multiple formats."""
    week_name = Path(week_file).stem
    
    # Load and generate schedule
    troops = load_troops_from_json(week_file)
    scheduler = ConstrainedScheduler(troops, get_all_activities())
    schedule = scheduler.schedule_all()
    
    # Create exports directory
    Path("exports").mkdir(exist_ok=True)
    
    # Export to requested formats
    if 'csv' in formats:
        export_to_csv(schedule, troops, f"exports/{week_name}.csv")
    
    if 'excel' in formats:
        export_to_excel(schedule, troops, f"exports/{week_name}.xlsx")
    
    if 'pdf' in formats:
        export_to_pdf(schedule, troops, f"exports/{week_name}.pdf")

if __name__ == "__main__":
    import glob
    import sys
    
    # Export all weeks
    week_files = glob.glob("tc_week*.json") + glob.glob("voyageur_week*.json")
    
    formats = sys.argv[1:] if len(sys.argv) > 1 else ['csv', 'excel', 'pdf']
    
    print(f"Exporting schedules to formats: {', '.join(formats)}\n")
    
    for week_file in sorted(week_files):
        print(f"\nExporting {week_file}...")
        export_schedule(week_file, formats)
    
    print("\n[OK] All exports complete!")
